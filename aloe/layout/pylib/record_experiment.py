from __future__ import division
import os
import numpy as np
import subprocess
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import aloe.config.config as cfg
import aloe.layout.pnr.usr_ip as ip

def read_xlsx(workbook, sheet):
    wb = load_workbook(workbook)
    sh = wb[sheet]
    df = pd.DataFrame(sh.values)
    # Set first row as column label
    df.columns = df.iloc[0]
    # Drop first row from data
    df.drop([0], inplace=True)
    # Reset row index to start from 0
    df = df.reset_index(drop=True)
    return df

def write_xlsx(workbook, experiment):

    tpnr = read_time(os.path.join(ip.pnr_dir, 'time_pnr.txt'))
    tstrmout = read_time(os.path.join(ip.pnr_dir, 'time_strout_pnr.txt')) - tpnr/float(ip.ngen)
    t_per_gen, t_per_gen_std = cal_generation_times(ip.output_dir, output_type='perf', INCL=False)
    version = get_git_tag()
    #TODO: not use dictionary since it is not ordered...
    dict_variables = {
        # Configuration
        'lib_name': [cfg.lib_name],
        'blk_name': [cfg.blk_name],
        'ncpu': [cfg.ncpu],
        'group': [cfg.group],
        'lpins': [list2str(cfg.lpins)],
        'rpins': [list2str(cfg.rpins)],
        'tpins': [list2str(cfg.tpins)],
        'bpins': [list2str(cfg.bpins)],
        # User Input File (some are auto-generated)
        'aspect_ratio': [ip.aspect_ratio],
        'area_util': [ip.area_util],
        'pin_file': [ip.pin_file],
        'top_layer': [ip.top_layer],
        'obj':[list2str(ip.perf)],
        'nout': [ip.nout],
        'pop_size': [ip.pop_size],
        'ngen': [ip.ngen],
        'cx_prob': [ip.cx_prob],
        'mut_prob': [ip.mut_prob],
        'eta': [ip.eta],
        'tpnr': [tpnr],
        't_per_gen_mean': [t_per_gen],
        't_per_gen_std': [t_per_gen_std],
        'tstrmout': [tstrmout],
        'version': [version]
    }
    df = pd.DataFrame(dict_variables)

    if os.path.exists(workbook):
        wb = load_workbook(workbook)
        ws=wb.create_sheet(title=experiment)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = experiment
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(workbook)

def read_time(ftime):
    '''
    ftime contains a single integer, unit second
    '''
    with open(ftime, 'r') as fin:
        return int(fin.readline().rstrip())

def cal_generation_times(output_dir, output_type='perf', INCL=False, OUT_MEAN=True, OUT_STD=True):
    '''
    output_type: output type
    incl: include the last file or not. Usually the last file is one that contains
          details for the streamed files.
    '''
    cmd = 'ls --time-style=+%s -ltr ' + os.path.join(output_dir, output_type)+' | awk \'{print ($6)}\''
    times_str = subprocess.check_output(cmd, shell=True)
    times = map(int, times_str.rstrip().split())
    idx_upper = len(times) if INCL else len(times)-1
    dtimes = np.zeros(idx_upper-1)
    for idx in range(1, idx_upper):
        dtimes[idx-1] = times[idx] - times[idx-1]

    mean = np.average(dtimes)
    std = np.std(dtimes)
    if OUT_MEAN & OUT_STD:
    	return mean, std
    elif OUT_MEAN & ~OUT_STD:
    	return mean
    elif ~OUT_MEAN & OUT_STD:
    	return std
    else:
    	return None

def get_git_tag():
    '''
    returns the git tag version. 
    example: release_0p3-19-g1c245ab
             * tag name: release_0p3
             * commits away from tag point: 19
             * short SHA1: 1c245ab
    '''
    return subprocess.check_output(['git','describe','--tags']).rstrip()
    #return subprocess.call(['git','describe','--tags'])

def copy_files(source, destination):
    if os.path.isfile(source):
        subprocess.call(['cp', source, destination])
    elif os.path.isdir(source):
        subprocess.call(['cp', '-r', source, destination])
    else:
        print 'ERROR: source is {}'.format(source)

def list2str(lis):
    if len(lis)>0:
        return ','.join(map(str, lis))
    else:
        return None

if __name__ == '__main__':

    name_exp = 'ro_top_6'
    dir_exp = '/r5sim/pohsuan/ro_top/experiments'

    # Copy
    dir_dest = os.path.join(dir_exp, name_exp)
    if os.path.exists(dir_dest):
        print 'WARN: {} Already exists. Nothing is copied'.format(dir_dest)
    else:
        copy_files(os.path.join(ip.lay_dir, 'output'), dir_dest)
        copy_files(os.path.join(ip.pnr_dir, 'norm_arr.npy'), os.path.join(dir_dest, 'perf', 'norm_arr.npy'))
        copy_files(os.path.join(ip.lay_dir, 'engine', 'expr_hof'), os.path.join(dir_dest, 'expr'))
        # Write excel files
        fxlsx = os.path.join(dir_exp, 'results.xlsx')
        write_xlsx(fxlsx, name_exp)

